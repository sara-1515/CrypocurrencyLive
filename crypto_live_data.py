import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import time

# Constants
API_URL = "https://api.coingecko.com/api/v3/coins/markets"
PARAMS = {
    'vs_currency': 'usd',
    'order': 'market_cap_desc',
    'per_page': 50,
    'page': 1,
    'sparkline': False
}
UPDATE_INTERVAL = 300  # Update every 5 minutes (300 seconds)

# Functions
def fetch_cryptocurrency_data():
    """Fetch live cryptocurrency data from the API."""
    response = requests.get(API_URL, params=PARAMS)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: Unable to fetch data (status code {response.status_code})")
        return []

def analyze_data(data):
    """Perform basic analysis on cryptocurrency data."""
    df = pd.DataFrame(data)
    df['price_change_percentage_24h'] = pd.to_numeric(df['price_change_percentage_24h'], errors='coerce')

    top_5_by_market_cap = df.nlargest(5, 'market_cap')[['name', 'symbol', 'market_cap']]
    average_price = df['current_price'].mean()
    highest_24h_change = df.loc[df['price_change_percentage_24h'].idxmax()]
    lowest_24h_change = df.loc[df['price_change_percentage_24h'].idxmin()]

    analysis = {
        "Top 5 by Market Cap": top_5_by_market_cap,
        "Average Price": average_price,
        "Highest 24h Change": highest_24h_change,
        "Lowest 24h Change": lowest_24h_change
    }

    return analysis


def write_to_excel(data, workbook):
    """Write the cryptocurrency data to an Excel sheet with styled headers."""
    sheet = workbook.active
    df = pd.DataFrame(data)

    # Write DataFrame to Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            # Convert complex objects (e.g., dicts) to string
            if isinstance(value, (dict, list)):
                value = str(value)
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)

            # Style the header row (first row)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")  # Bold and white text
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', ...)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        sheet.column_dimensions[col_letter].width = adjusted_width



def main():
    """Main function to fetch, analyze, and write data."""
    workbook = Workbook()

    while True:
        data = fetch_cryptocurrency_data()
        if not data:
            break

        # Data Analysis
        analysis = analyze_data(data)
        print("Analysis:")
        print(analysis)

        # Write to Excel
        write_to_excel(data, workbook)

        # Save the Excel file
        workbook.save("live_cryptocurrency_data.xlsx")

        print("Data updated in Excel. Next update in 5 minutes...")
        time.sleep(UPDATE_INTERVAL)

if __name__ == "__main__":
    main()
